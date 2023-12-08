import os
import sys
from configparser import ConfigParser

from PyQt5 import QtCore, QtGui, QtWidgets
import tkinter as tk
from tkinter import filedialog

from PyQt5.QtCore import pyqtSignal
from PyQt5.QtWidgets import QPlainTextEdit, QDialog, QMessageBox, QComboBox

from openpyxl import load_workbook


class Ui_Form(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(800, 650)

        self.comboBox = QtWidgets.QComboBox(Dialog)
        self.comboBox.setGeometry(QtCore.QRect(530, 40, 181, 41))
        self.comboBox.setObjectName("comboBox")
        # self.comboBox.setEditable(True)
        self.comboBox.setInsertPolicy(QComboBox.NoInsert)

        config = ConfigParser()
        config.read('E:\\comboBox.ini')

        for section in config.sections():
            if section != "SSSSSSS":
                for key, value in config.items(section):
                    self.comboBox.addItem(value)

        # excel_file_path = 'E:\\機能一覧.xlsx'
        # workbook = load_workbook(excel_file_path)
        #
        # sheet = workbook["Sheet1"]
        #
        # for row in sheet.iter_rows(min_row=2, values_only=True):
        #     if row[0]:
        #         self.comboBox.addItem(str(row[1]) + " : " + str(row[2]))
        #
        # sheet.sheet_view.zoomScale = 85
        # sheet.sheet_view.selection[0].activeCell = 'A1'
        # sheet.sheet_view.selection[0].sqref = 'A1'
        # # for row in sheet.iter_rows():
        # #     # row.reset_dimensions()
        # #     row.row_dimensions.height = 25
        # #     # for cell in row:
        # #     #     cell.row_dimensions.height = 25
        # # workbook.active.index = 0
        # # workbook.active = workbook["Sheet1"]
        # workbook.active = sheet
        # workbook.save(excel_file_path)

        self.plainTextEdit = QtWidgets.QLineEdit(Dialog)
        self.plainTextEdit.setGeometry(QtCore.QRect(60, 40, 271, 31))
        self.plainTextEdit.setObjectName("plainTextEdit")
        self.plainTextEdit_2 = QtWidgets.QLineEdit(Dialog)
        self.plainTextEdit_2.setGeometry(QtCore.QRect(60, 90, 271, 31))
        self.plainTextEdit_2.setObjectName("plainTextEdit_2")
        self.pushButton = QtWidgets.QPushButton(Dialog)
        self.pushButton.setGeometry(350, 40, 91, 31)
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.on_pushButton_click)
        self.pushButton_2 = QtWidgets.QPushButton(Dialog)
        self.pushButton_2.setGeometry(350, 90, 91, 31)
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.on_pushButton_click)
        self.tabWidget = QtWidgets.QTabWidget(Dialog)
        self.tabWidget.setGeometry(QtCore.QRect(10, 230, 771, 271))
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.listWidget = QtWidgets.QListWidget(self.tab)
        self.listWidget.setGeometry(QtCore.QRect(0, 0, 751, 231))
        self.listWidget.setObjectName("listWidget")
        # self.horizontalScrollBar = QtWidgets.QScrollBar(self.tab)
        # self.horizontalScrollBar.setGeometry(QtCore.QRect(0, 230, 751, 16))
        # self.horizontalScrollBar.setOrientation(QtCore.Qt.Horizontal)
        # self.horizontalScrollBar.setObjectName("horizontalScrollBar")
        # self.verticalScrollBar = QtWidgets.QScrollBar(self.tab)
        # self.verticalScrollBar.setGeometry(QtCore.QRect(750, 0, 16, 241))
        # self.verticalScrollBar.setOrientation(QtCore.Qt.Vertical)
        # self.verticalScrollBar.setObjectName("verticalScrollBar")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.tableWidget = QtWidgets.QTableWidget(self.tab_2)
        self.tableWidget.setGeometry(QtCore.QRect(0, 0, 751, 231))
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        # self.horizontalScrollBar_2 = QtWidgets.QScrollBar(self.tab_2)
        # self.horizontalScrollBar_2.setGeometry(QtCore.QRect(0, 230, 751, 16))
        # self.horizontalScrollBar_2.setOrientation(QtCore.Qt.Horizontal)
        # self.horizontalScrollBar_2.setObjectName("horizontalScrollBar_2")
        # self.verticalScrollBar_2 = QtWidgets.QScrollBar(self.tab_2)
        # self.verticalScrollBar_2.setGeometry(QtCore.QRect(750, 0, 16, 241))
        # self.verticalScrollBar_2.setOrientation(QtCore.Qt.Vertical)
        # self.verticalScrollBar_2.setObjectName("verticalScrollBar_2")
        self.tabWidget.addTab(self.tab_2, "")
        self.pushButton_3 = QtWidgets.QPushButton(Dialog)
        self.pushButton_3.setGeometry(60, 140, 111, 41)
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QtWidgets.QPushButton(Dialog)
        self.pushButton_4.setGeometry(220, 140, 111, 41)
        self.pushButton_4.setObjectName("pushButton_4")
        # self.label = QtWidgets.QLabel(Dialog)
        # self.label.setGeometry(QtCore.QRect(530, 40, 181, 41))
        # self.label.setObjectName("label")
        self.label_1 = QtWidgets.QLabel(Dialog)
        self.label_1.setGeometry(QtCore.QRect(10, 520, 771, 41))
        self.label_1.setObjectName("label_1")

        self.retranslateUi(Dialog)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Dialog)

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Dialog"))
        self.pushButton.setText(_translate("Dialog", "PushButton"))
        self.pushButton_2.setText(_translate("Dialog", "PushButton"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("Dialog", "Tab_1"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("Dialog", "Tab_2"))
        self.pushButton_3.setText(_translate("Dialog", "PushButton"))
        self.pushButton_4.setText(_translate("Dialog", "PushButton"))
        # self.label.setText(_translate("Dialog", "TextLabel"))
        self.label_1.setText(_translate("Dialog", "TextLabel"))

    def on_pushButton_click(self):
        # folder_path = filedialog.askdirectory()
        # if folder_path:
        #     self.plainTextEdit.clear()
        #     self.plainTextEdit.setText(folder_path)
        #     # show_files_in_tree(folder_path)
        QMessageBox.information(self, 'info', "test in UI！")

    def on_focus_lost(self):
        if self.plainTextEdit.text().replace(" ", "") == "":
            return
        else:
            if os.path.exists(self.plainTextEdit.text()):
                self.pushButton_3.config(state=tk.NORMAL)
                # self.label.config(text="")
            else:
                self.pushButton_3.config(state=tk.DISABLED)
                # self.label.config(text=f"ファイル['{self.plainTextEdit.text()}']はシステムに存在しません。",
                #                   fg="red", font=('Arial', 13))
                self.plainTextEdit.focus_set()


class MyDialog(QtWidgets.QDialog, Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.plainTextEdit.installEventFilter(self)
        # self.pushButton.installEventFilter(self)
        self.pushButton_2.installEventFilter(self)
        self.pushButton_3.installEventFilter(self)
        self.pushButton_4.installEventFilter(self)
        self.comboBox.installEventFilter(self)

    def eventFilter(self, source, event):
        if source == self.plainTextEdit and event.type() == QtCore.QEvent.Type.FocusOut:
            directory_path = self.plainTextEdit.text()

            if not os.path.exists(directory_path):
                self.pushButton_3.setEnabled(False)
            else:
                self.pushButton_3.setEnabled(True)

        if source == self.pushButton and event.type() == QtCore.QEvent.Type.MouseButtonRelease:
            folder_path = filedialog.askdirectory()
            if folder_path:
                self.pushButton_3.setEnabled(True)
                self.plainTextEdit.clear()
                self.plainTextEdit.setText(folder_path)
                self.plainTextEdit.setReadOnly(True)
                self.listWidget.clear()
                listWidgetForShow = []
                for root, dirs, files in os.walk(folder_path):
                    for file in files:
                        if file.endswith(".java"):
                            # listWidgetForShow.append(root + "/" + file)
                            listWidgetForShow.append(os.path.join(root.replace("/", "\\"), file))
                self.listWidget.addItems(listWidgetForShow)
                self.listWidget.scrollToBottom()

        if source == self.pushButton_2 and event.type() == QtCore.QEvent.Type.MouseButtonRelease:
            QMessageBox.information(self, 'info', "test in LOGIC！")

        if source == self.pushButton_4 and event.type() == QtCore.QEvent.Type.MouseButtonRelease:
            app.exit()

        if source == self.pushButton_3 and event.type() == QtCore.QEvent.Type.MouseButtonRelease:
            if self.plainTextEdit != "" and self.plainTextEdit_2 != "":
                os.makedirs(os.path.join(self.plainTextEdit.text(), self.plainTextEdit_2.text(), "LOG"), exist_ok=True)
                os.makedirs(os.path.join(self.plainTextEdit.text(), self.plainTextEdit_2.text(), "LOG", "JAVA側"),
                            exist_ok=True)
                os.makedirs(os.path.join(self.plainTextEdit.text(), self.plainTextEdit_2.text(), "LOG", "HOST側"),
                            exist_ok=True)
                os.makedirs(os.path.join(self.plainTextEdit.text(), self.plainTextEdit_2.text(), "カバレッジ"),
                            exist_ok=True)
                os.makedirs(os.path.join(self.plainTextEdit.text(), self.plainTextEdit_2.text(), "現新比較"),
                            exist_ok=True)
                os.makedirs(os.path.join(self.plainTextEdit.text(), self.plainTextEdit_2.text(), "現新比較", "JAVA側"),
                            exist_ok=True)
                os.makedirs(os.path.join(self.plainTextEdit.text(), self.plainTextEdit_2.text(), "現新比較", "HOST側"),
                            exist_ok=True)
            QMessageBox.information(self, '完成', "フォルダーを創建成功！")

        if source == self.comboBox and event.type() == QtCore.QEvent.Type.KeyRelease:
            value = self.comboBox.itemText()

        return super().eventFilter(source, event)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')
    # MainWindow = QtWidgets.QMainWindow()
    # ui = Ui_Form()
    # ui.setupUi(MainWindow)
    # MainWindow.show()
    dialog = MyDialog()
    dialog.show()
    app.exec_()
