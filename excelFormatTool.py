import os
import subprocess
import sys
import time
from tkinter import filedialog

import pyautogui
from openpyxl import load_workbook
from PyQt5.QtCore import QDateTime, Qt, QTimer
from PyQt5.QtWidgets import QApplication, QGroupBox, QLabel, QListWidget, QMessageBox, QProgressBar, QPushButton, \
	QVBoxLayout, QWidget


class MyWindow(QWidget):
	def __init__(self):
		super().__init__()

		self.process_count = None
		self.source_file_path = None
		self.files_list = None
		self.timer_show_files = None
		self.button_exe = None
		self.folder_path = None
		self.timer = None
		self.label = None
		self.process_bar = None
		self.list_widget = None
		self.group_layout = None
		self.group_box = None
		self.button = None
		self.main_layout = None
		self.initUI()

	def initUI(self):
		self.main_layout = QVBoxLayout(self)
		self.button = QPushButton('EXCELファイルパスを選んでください。')
		self.main_layout.addWidget(self.button)
		self.button.clicked.connect(self.on_button_click)
		self.group_box = QGroupBox('EXCELファイル一覧')
		self.group_layout = QVBoxLayout(self.group_box)
		self.list_widget = QListWidget()
		self.group_layout.addWidget(self.list_widget)
		self.main_layout.addLayout(self.group_layout)
		self.process_bar = QProgressBar()
		self.process_bar.setMinimum(0)
		self.process_bar.setMaximum(100)
		self.group_layout.addWidget(self.process_bar)
		self.main_layout.addWidget(self.group_box)

		self.button_exe = QPushButton('EXCELファイルを転換する。')
		self.main_layout.addWidget(self.button_exe)
		self.button_exe.clicked.connect(self.on_button_exe_click)

		self.label = QLabel("")
		self.main_layout.addWidget(self.label)

		self.setGeometry(100, 100, 600, 300)
		self.setWindowTitle("excelFormatTool")

		self.timer_init()

	def timer_init(self):
		self.timer = QTimer()
		self.timer.timeout.connect(self.update_datetime_and_progress)
		self.timer.start(1000)

	def update_datetime_and_progress(self):
		current_datetime = QDateTime.currentDateTime().toString(Qt.ISODate)
		self.label.setText(f'ローカル　タイム：{current_datetime}')

	def on_button_click(self):
		self.folder_path = filedialog.askdirectory()
		self.show_files_in_tree()

	def show_files_in_tree(self):
		print(f'show_files_in_tree start...')
		self.process_count = 0
		# self.timer_show_files = QTimer()
		# self.timer_show_files.timeout.connect(self.update_progress)
		# self.timer_show_files.start(1000)
		self.files_list = os.listdir(self.folder_path)
		print(f'folder_path : {self.folder_path}')
		self.list_widget.clear()
		for root, _, files in os.walk(self.folder_path):
			for filename in files:
				self.source_file_path = os.path.join(root, filename)
				if os.path.isfile(self.source_file_path) and self.source_file_path.endswith(".xlsx"):
					self.process_count += 1
					self.update_progress()

	def update_progress(self):
		self.source_file_path = self.source_file_path.replace("/", "\\")
		self.list_widget.addItem(self.source_file_path)
		self.list_widget.scrollToBottom()

	def on_button_exe_click(self):
		for row in range(self.list_widget.count()):
			item = self.list_widget.item(row)
			self.file = item.text()
			print(f'file : {self.file}')

			self.open_excel_file()
			self.activate_excel_window()
			self.simulate_scroll()
		# workbook = load_workbook(file)
		# all_sheets = workbook.sheetnames
		#
		# for sheet_name in all_sheets:
		# 	sheet = workbook[sheet_name]
		# 	sheet.sheet_view.zoomScale = 85
		# 	sheet.sheet_properties.tabColor = None
		# 	sheet.sheet_view.selection[0].activeCell = 'A1'
		# 	sheet.sheet_view.selection[0].sqref = 'A1'
		# 	# sheet.views.sheetView[0].topLeftCell = "A1"
		# 	pyautogui.press('up')
		# 	time.sleep(1)
		# 	pyautogui.hotkey('ctrl', 'home')
		# 	start_cell = (100, 100)
		# 	data_to_fill = "Hello,Excel"
		# 	pyautogui.click(start_cell)
		# 	pyautogui.typewrite(data_to_fill, interval=0.1)
		# workbook.active = 0
		# workbook.save(file)

		QMessageBox.information(self, "Task Completed", "Process completed successfully!")

	def open_excel_file(self):
		subprocess.Popen(['start', 'excel', self.file], shell=True)
		time.sleep(2)

	def activate_excel_window(self):
		print(f'activate_excel_window start...')
		# excel_window_title = 'Microsoft Excel'
		# pyautogui.click(pyautogui.locateCenterOnScreen(excel_window_title + '.png'))
		title_bar_position = pyautogui.locateOnScreen('excel_title_bar.png')
		print(f'title_bar_position : {title_bar_position}')
		if title_bar_position:
			title_bar_center = pyautogui.center(title_bar_position)
			pyautogui.click(title_bar_center[0], title_bar_center[1])

	def simulate_scroll(self):
		scrollbar_position = (100, 200)
		pyautogui.click(scrollbar_position[0], scrollbar_position[1])


if __name__ == '__main__':
	app = QApplication(sys.argv)
	window = MyWindow()
	window.show()
	sys.exit(app.exec_())
